import openpyxl
from openpyxl import load_workbook

def check_all_columns():
    """모든 열의 헤더와 데이터 확인"""
    
    file_path = '골구조도.xlsx'
    wb = load_workbook(file_path, data_only=False)
    ws = wb.active
    
    print("=" * 80)
    print("전체 열 헤더 확인 (1~10행)")
    print("=" * 80)
    
    # 각 열의 헤더 확인
    for col_idx in range(1, ws.max_column + 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        col_headers = []
        
        for row_idx in range(1, 11):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None or cell.data_type == 'f':
                if cell.data_type == 'f':
                    col_headers.append(f"행{row_idx}: {cell.value}")
                else:
                    value_str = str(cell.value)
                    if len(value_str) > 30:
                        value_str = value_str[:30] + "..."
                    col_headers.append(f"행{row_idx}: {value_str}")
        
        if col_headers:
            print(f"\n{col_letter}열:")
            for header in col_headers:
                print(f"  {header}")
    
    print("\n" + "=" * 80)
    print("L, M, N열 전체 데이터 확인")
    print("=" * 80)
    
    # L, M, N열의 모든 데이터 확인
    for col_idx in [12, 13, 14]:  # L, M, N열
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        print(f"\n{col_letter}열 전체 데이터:")
        
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None or cell.data_type == 'f':
                if cell.data_type == 'f':
                    print(f"  {col_letter}{row_idx}: {cell.value}")
                else:
                    print(f"  {col_letter}{row_idx}: {cell.value}")

if __name__ == '__main__':
    check_all_columns()

