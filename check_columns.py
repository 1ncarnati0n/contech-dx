import openpyxl
from openpyxl import load_workbook

def check_lmn_columns():
    """L, M, N열의 데이터 확인"""
    
    file_path = '골구조도.xlsx'
    wb = load_workbook(file_path, data_only=False)
    ws = wb.active
    
    print("=" * 80)
    print("L, M, N열 데이터 확인")
    print("=" * 80)
    
    # L, M, N열의 모든 데이터 확인
    for row_idx in range(1, ws.max_row + 1):
        l_cell = ws.cell(row=row_idx, column=12)  # L열
        m_cell = ws.cell(row=row_idx, column=13)  # M열
        n_cell = ws.cell(row=row_idx, column=14)  # N열
        
        row_data = []
        if l_cell.value is not None or l_cell.data_type == 'f':
            row_data.append(f"L{row_idx}: {l_cell.value} (타입: {l_cell.data_type})")
        if m_cell.value is not None or m_cell.data_type == 'f':
            row_data.append(f"M{row_idx}: {m_cell.value} (타입: {m_cell.data_type})")
        if n_cell.value is not None or n_cell.data_type == 'f':
            row_data.append(f"N{row_idx}: {n_cell.value} (타입: {n_cell.data_type})")
        
        if row_data:
            print(f"\n행 {row_idx}:")
            for item in row_data:
                print(f"  {item}")
    
    # 헤더 행 확인 (1~10행)
    print("\n" + "=" * 80)
    print("헤더 영역 확인 (1~10행)")
    print("=" * 80)
    
    for row_idx in range(1, 11):
        row_data = []
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None or cell.data_type == 'f':
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                if cell.data_type == 'f':
                    row_data.append(f"{col_letter}{row_idx}: {cell.value}")
                else:
                    value_str = str(cell.value)
                    if len(value_str) > 50:
                        value_str = value_str[:50] + "..."
                    row_data.append(f"{col_letter}{row_idx}: {value_str}")
        if row_data:
            print(f"\n행 {row_idx}:")
            for item in row_data:
                print(f"  {item}")

if __name__ == '__main__':
    check_lmn_columns()

