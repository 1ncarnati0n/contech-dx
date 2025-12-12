import openpyxl
from openpyxl import load_workbook

def check_material_data():
    """형틀, 철근, 콘크리트 물량 데이터 확인"""
    
    file_path = '골구조도.xlsx'
    wb = load_workbook(file_path, data_only=False)
    ws = wb.active
    
    print("=" * 80)
    print("전체 시트 데이터 확인 (형틀, 철근, 콘크리트 관련)")
    print("=" * 80)
    
    # 모든 셀에서 형틀, 철근, 콘크리트 관련 키워드 검색
    keywords = ['형틀', '갱폼', '알폼', '철근', '콘크리트', '물량', 'M2', 'TON', 'M3']
    
    found_cells = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                value_str = str(cell.value).upper()
                for keyword in keywords:
                    if keyword.upper() in value_str:
                        found_cells.append({
                            'cell': cell.coordinate,
                            'value': str(cell.value),
                            'row': cell.row,
                            'col': cell.column
                        })
                        break
    
    if found_cells:
        print("\n형틀/철근/콘크리트 관련 셀:")
        for cell_info in found_cells:
            print(f"  {cell_info['cell']}: {cell_info['value']}")
    else:
        print("\n형틀/철근/콘크리트 관련 키워드를 찾을 수 없습니다.")
    
    # 전체 시트 구조 확인 (모든 행, 모든 열)
    print("\n" + "=" * 80)
    print("전체 시트 구조 (모든 행의 모든 열)")
    print("=" * 80)
    
    for row_idx in range(1, ws.max_row + 1):
        row_data = []
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None or cell.data_type == 'f':
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                if cell.data_type == 'f':
                    row_data.append(f"{col_letter}: {cell.value}")
                else:
                    value_str = str(cell.value)
                    if len(value_str) > 20:
                        value_str = value_str[:20] + "..."
                    row_data.append(f"{col_letter}: {value_str}")
        
        if row_data:
            print(f"\n행 {row_idx}:")
            for item in row_data:
                print(f"  {item}")

if __name__ == '__main__':
    check_material_data()

